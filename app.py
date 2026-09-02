"""
Assignment Design Assistant
============================
A small Flask web app that lets students fill in, save, and submit the
sections required by an "Applications Development" style assignment brief
(Design Document -> Build -> Testing -> Evaluation) online, instead of in a
blank Word document.

Responses are stored in PostgreSQL (see schema.sql / models.py) and can be
exported to Excel by a tutor/admin for marking or record-keeping.

Run locally:
    export DATABASE_URL=postgresql://user:pass@localhost:5432/assignment_writer
    export ADMIN_PASSWORD=change-me
    flask --app app run --debug
"""
import os
from datetime import datetime, date
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for, flash,
    send_file, session, abort
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from extensions import db
from models import AssignmentResponse

# ---------------------------------------------------------------------------
# App & database configuration
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-secret-change-me")

db_url = os.environ.get("DATABASE_URL", "sqlite:///local_dev.db")
# Render/Heroku-style URLs start with postgres:// - SQLAlchemy 1.4+ wants postgresql://
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {"pool_pre_ping": True}

db.init_app(app)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")

APP_TYPES = ["Mobile App", "Game App", "Business App", "Web App", "Other"]
DEV_ENVIRONMENTS = [
    "Python (IDLE)", "Python (VS Code)", "MIT App Inventor",
    "Scratch (Advanced)", "Java (BlueJ)", "Java (Eclipse)", "Other",
]


def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    """The blank/new submission form."""
    return render_template(
        "form.html",
        response=None,
        field_groups=AssignmentResponse.FIELD_GROUPS,
        app_types=APP_TYPES,
        dev_environments=DEV_ENVIRONMENTS,
        today=date.today().isoformat(),
    )


@app.route("/edit/<int:response_id>")
def edit(response_id):
    """Re-open a previously saved draft/submission for editing."""
    response = AssignmentResponse.query.get_or_404(response_id)
    return render_template(
        "form.html",
        response=response,
        field_groups=AssignmentResponse.FIELD_GROUPS,
        app_types=APP_TYPES,
        dev_environments=DEV_ENVIRONMENTS,
        today=date.today().isoformat(),
    )


@app.route("/submit", methods=["POST"])
def submit():
    """Create or update a response. The 'action' field decides draft vs final."""
    response_id = request.form.get("response_id")
    action = request.form.get("action", "draft")  # 'draft' or 'submit'

    if response_id:
        record = AssignmentResponse.query.get_or_404(int(response_id))
    else:
        record = AssignmentResponse()
        db.session.add(record)

    # Cover sheet / basics
    record.student_usn = request.form.get("student_usn", "").strip()
    record.student_name = request.form.get("student_name", "").strip()
    record.tutor_name = request.form.get("tutor_name", "").strip()
    record.unit_code = request.form.get("unit_code", "H6S9 46").strip()
    record.unit_title = request.form.get("unit_title", "Computing: Applications Development").strip()
    record.date_due = parse_date(request.form.get("date_due"))
    record.date_submitted = parse_date(request.form.get("date_submitted"))
    record.app_name = request.form.get("app_name", "").strip()
    record.app_type = request.form.get("app_type", "").strip()
    record.dev_environment = request.form.get("dev_environment", "").strip()

    if not record.student_usn or not record.student_name:
        flash("Student USN and Student Name are required.", "error")
        return redirect(request.referrer or url_for("index"))

    # All the task fields
    for group in AssignmentResponse.FIELD_GROUPS:
        for field_name, _label, _hint, _marks in group["fields"]:
            setattr(record, field_name, request.form.get(field_name, "").strip())

    record.status = "submitted" if action == "submit" else "draft"

    db.session.commit()
    flash(
        "Final submission saved." if record.status == "submitted" else "Draft saved.",
        "success",
    )
    return redirect(url_for("success", response_id=record.id))


@app.route("/success/<int:response_id>")
def success(response_id):
    record = AssignmentResponse.query.get_or_404(response_id)
    return render_template("success.html", response=record)


# ---------------------------------------------------------------------------
# Admin / tutor area — view all responses, export to Excel
# ---------------------------------------------------------------------------
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST":
        if request.form.get("password") == ADMIN_PASSWORD:
            session["is_admin"] = True
        else:
            flash("Incorrect password.", "error")
            return redirect(url_for("admin"))

    if not session.get("is_admin"):
        return render_template("admin_login.html")

    responses = AssignmentResponse.query.order_by(AssignmentResponse.created_at.desc()).all()
    return render_template("admin.html", responses=responses)


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin"))


@app.route("/admin/export.xlsx")
def export_excel():
    if not session.get("is_admin"):
        abort(403)

    responses = AssignmentResponse.query.order_by(AssignmentResponse.student_usn).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Assignment Responses"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    if responses:
        columns = list(responses[0].to_export_row().keys())
    else:
        columns = list(AssignmentResponse().to_export_row().keys())

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 28

    for row_idx, record in enumerate(responses, start=2):
        row_data = record.to_export_row()
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"assignment_responses_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/response/<int:response_id>")
def admin_view_response(response_id):
    if not session.get("is_admin"):
        abort(403)
    record = AssignmentResponse.query.get_or_404(response_id)
    return render_template("success.html", response=record, admin_view=True)


# ---------------------------------------------------------------------------
# Health check (handy for Render)
# ---------------------------------------------------------------------------
@app.route("/healthz")
def healthz():
    return {"status": "ok"}, 200


with app.app_context():
    db.create_all()


if __name__ == "__main__":
    app.run(debug=True)
